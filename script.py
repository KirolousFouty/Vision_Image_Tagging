# Import necessary libraries
from transformers import AutoProcessor, AutoModelForCausalLM
from PIL import Image, UnidentifiedImageError
import requests
import copy
import shutil
import os
import re
from googletrans import Translator
import openpyxl
from openpyxl import Workbook
import google.generativeai as genai
from google.colab import userdata
from datetime import datetime

# Import API key from config.py
from config import GEMINI_API_KEY

# Load the model and processor
model_id = 'microsoft/Florence-2-large'
try:
    model = AutoModelForCausalLM.from_pretrained(model_id, trust_remote_code=True).eval().cuda()
    processor = AutoProcessor.from_pretrained(model_id, trust_remote_code=True)
except Exception as e:
    raise Exception(f"Failed to load model or processor: {str(e)}")

# Define the prediction function
def run_example(task_prompt, text_input=None, image=None):
    try:
        inputs = processor(text=task_prompt + (text_input if text_input else ""), images=image, return_tensors="pt")
        generated_ids = model.generate(
            input_ids=inputs["input_ids"].cuda(),
            pixel_values=inputs["pixel_values"].cuda(),
            max_new_tokens=1024,
            early_stopping=False,
            do_sample=False,
            num_beams=3,
        )
        generated_text = processor.batch_decode(generated_ids, skip_special_tokens=False)[0]
        parsed_answer = processor.post_process_generation(
            generated_text,
            task=task_prompt,
            image_size=(image.width, image.height)
        )
        return parsed_answer
    except Exception as e:
        raise Exception(f"Error during model inference: {str(e)}")

# Configure Gemini API
try:
    genai.configure(api_key=GEMINI_API_KEY)
except Exception as e:
    raise Exception(f"Error configuring Gemini API: {str(e)}")

# Function to extract metadata from filename
def extract_metadata_from_filename(filename):
    try:
        match = re.match(r"(?P<SOURCE_TITLE>.+)_(?P<DATE>\d{4})_(?P<SOURCE_NUMBER>iss\d+)_(?P<PAGE>Page\d+).jpg", filename)
        if match:
            metadata = match.groupdict()
            metadata["SOURCE_NUMBER"] = metadata["SOURCE_NUMBER"].replace("iss", "")
            metadata["PAGE_NUMBER"] = metadata.pop("PAGE").replace("Page", "")
            return metadata
        else:
            raise ValueError(f"Filename {filename} does not match the expected pattern")
    except Exception as e:
        raise Exception(f"Error extracting metadata from filename: {str(e)}")

# Function to translate text to Arabic
def translate_to_arabic(text):
    try:
        translator = Translator()
        translated = translator.translate(text, dest='ar')
        return translated.text
    except Exception as e:
        raise Exception(f"Translation error: {str(e)}")

# Function to process an image and extract metadata
def process_image(image_path, output_dir):
    try:
        print(f"Processing {image_path}")
        image = Image.open(image_path)
    except FileNotFoundError:
        raise Exception(f"Image file {image_path} not found.")
    except UnidentifiedImageError:
        raise Exception(f"Image file {image_path} is not a valid image.")
    except Exception as e:
        raise Exception(f"Error opening image file: {str(e)}")

    try:
        # Get detailed caption from the image
        task_prompt = '<MORE_DETAILED_CAPTION>'
        results = run_example(task_prompt, image=image)
        alt_text = results[task_prompt]
        text_input = results[task_prompt]

        # Get phrase grounding from the caption
        task_prompt = '<CAPTION_TO_PHRASE_GROUNDING>'
        results = run_example(task_prompt, text_input, image=image)
        labels_list = ', '.join(results["<CAPTION_TO_PHRASE_GROUNDING>"]["labels"])

        # Prepare text manipulation prompt for Gemini
        text = f"I will provide you with a comma-separated list of words. Remove all articles (e.g., 'a', 'an', 'the'). Remove all pronouns. Identify and return only the most significant word from each value. Convert all words to their singular forms. Make all words lowercase. Remove any duplicate words. Return the output comma-separated. The list is as follows: {labels_list}"
        model = genai.GenerativeModel('gemini-pro')
        chat = model.start_chat(history=[])
        response = chat.send_message(text)

        # Split response into a list and sort
        def split_string_into_list(input_string):
            return [word.strip() for word in input_string.split(',')]

        result = sorted(set(split_string_into_list(response.text)))
    except Exception as e:
        raise Exception(f"Error processing image: {str(e)}")

    try:
        filename = os.path.basename(image_path)
        metadata = extract_metadata_from_filename(filename)

        # Generate title and additional metadata
        caption = translate_to_arabic(response.text)
        keywords = response.text
        title = f"{os.path.splitext(filename)[0]}__{keywords[:30].replace(' ', '_').replace(',', '')}"
        filesize = str(os.path.getsize(image_path))
        width, height = image.size
        dimensions = f"{height} x {width}"
        _, file_extension = os.path.splitext(filename)

        metadata.update({
            "FILENAME": filename,
            "TITLE": title,
            "DESCRIPTION": "AI-Generated: " + alt_text,
            "FILETYPE": file_extension,
            "FILESIZE": filesize,
            "DIMENSIONS": dimensions,
            "CAPTION": caption,
            "KEYWORDS": keywords
        })

        # Move processed image to output directory
        shutil.move(image_path, os.path.join(output_dir, filename))

        return metadata
    except Exception as e:
        raise Exception(f"Error updating metadata or moving image file: {str(e)}")

# Function to save metadata to an Excel file
def save_metadata_to_excel(all_metadata, output_path):
    try:
        wb = Workbook()
        ws = wb.active

        # Define and write headers
        headers = [
            "FILENAME", "SOURCE_TITLE", "SOURCE_NUMBER", "PAGE_NUMBER", "DATE",
            "TITLE", "DESCRIPTION", "FILETYPE", "FILESIZE", "DIMENSIONS",
            "CAPTION", "KEYWORDS"
        ]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        # Write metadata
        for row_idx, metadata in enumerate(all_metadata, start=2):
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=metadata[header])

        wb.save(output_path)
    except Exception as e:
        raise Exception(f"Error saving metadata to Excel: {str(e)}")

# Function to get the next run sequence number
def get_next_run_number(output_dir):
    try:
        subfolders = [f.name for f in os.scandir(output_dir) if f.is_dir()]
        run_numbers = [int(re.search(r'Run (\d+)', folder).group(1)) for folder in subfolders if re.search(r'Run (\d+)', folder)]
        return max(run_numbers, default=0) + 1
    except Exception as e:
        raise Exception(f"Error getting next run number: {str(e)}")

# Define directories
input_dir = "input/"
output_dir = "output/"

# Get next run number and create a subfolder
try:
    run_number = get_next_run_number(output_dir)
    run_datetime = datetime.now().strftime('%d %B %Y - %I:%M %p').upper()
    run_subfolder = f"Run {run_number} - {run_datetime}"
    output_subdir = os.path.join(output_dir, run_subfolder)

    os.makedirs(output_subdir, exist_ok=True)
    output_file = os.path.join(output_subdir, "metadata.xlsx")
except Exception as e:
    raise Exception(f"Error creating output subfolder: {str(e)}")

# Process all images in the input directory
all_metadata = []
try:
    for filename in os.listdir(input_dir):
        if filename.endswith(".jpg"):
            image_path = os.path.join(input_dir, filename)
            metadata = process_image(image_path, output_subdir)
            all_metadata.append(metadata)

    # Save metadata to an Excel file
    save_metadata_to_excel(all_metadata, output_file)
except Exception as e:
    raise Exception(f"Error processing images or saving metadata: {str(e)}")
